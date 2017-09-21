import sys
import numpy as np
from imp import reload
import time
import urllib.request
from bs4 import BeautifulSoup
from openpyxl import Workbook



reload(sys)

hds = [{'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36'}]
url = 'https://www.zhihu.com/topic/19606792/top-answers'


def book_spider(book_tag):
	page_num = 0
	book_list = []
	try_times = 0

	while(1):
		time.sleep(np.random.random()*5)

		plain_text = get_plain_text(url, hds)
		if plain_text == None:
			continue


		soup = BeautifulSoup(plain_text, 'html.parser')
		list_soup = soup.find('div', {'class':'mod book_list'})

		try_times += 1
		if list_soup == None and try_times < 200:
			continue
		elif list_soup == None or len(list_soup) <= 1:
			break

		for book_info in list_soup.findAll('dd'):
			title = book_info.find('a', {'class' : 'title'}).string.strip()
			desc = book_info.find('div', {'class' : 'desc'}).string.strip()
			desc_list = desc.split('/')
			book_url = book_info.find('a', {'class' : 'title'}).get('href')

			try:
				author_info = '作者/译者：' + '/'.join(desc_list[0:-3])
			except:
				author_info = '作者/译者：暂无'

			try:
				pub_info = '出版信息：' + '/'.join(desc_list[-3:])
			except:
				pub_info = '出版信息：暂无'

			try:
				rating = book_info.find('span', {'class': 'rating_nums'}).str.strip()
			except:
				rating = '0.0'

			try:
				people_num = get_people_num(book_url)
				people_num = people_num.strip('人评价')
			except:
				people_num = '0'

			book_list.append([title, rating, people_num, author_info, pub_info])
			try_times += 1
		page_num += 1
		print('Downloading Infomation From Page %d ' % page_num)
	return book_list


def get_plain_text(url, hds):
	try:
		req = urllib.request.Request(url, headers=hds[0])
		source_code = urllib.request.urlopen(req).read()
		plain_text = str(source_code)
	except urllib.error.HTTPError as e:
		print('has error')
		print(e)
	else:
		return plain_text


def get_people_num(url):
	try:
		req = urllib.Request(url, headers=hds[np.random.randint(0, len(hds))])
		source_code = urllib.request.urlopen(req).read()
		plain_text = str(source_code)
	except urllib.error.HTTPError as e:
		print(e)

	soup = BeautifulSoup(plain_text)
	people_num = soup.find('div', {'calss' : 'rating_sum'}).findAll('span')[1].str.strip()
	return people_num


def do_spider(book_tag_lists):
	book_lists = []
	for book_tag in book_tag_lists:
		book_list = book_spider(book_tag)
		book_list = sorted(book_list, key=lambda x:x[1], reverse=True)
		book_lists.append(book_list)

		return book_lists


def print_book_lists_excel(book_list, book_tag_lists):
	wb = Workbook(optimized_write=True)
	ws = []

	for i in range(len(book_tag_lists)):
		ws.append(wb.create_sheet(title=book_tag_lists[i]).decode())

	for i in range(len(book_tag_lists)):
		ws[i].append(['序号', '书名', '评分', '评价人数', '作者', '出版社'])
		count = 1
		for bl in book_lists[i]:
			ws[i].append([count, bl[0], float(bl[1]), int(bl[2]), bl[3], bl[4]])
			count += 1

	save_path = 'book_list'

	for i in range(len(book_tag_lists)):
		save_path += ('-' + book_tag_lists[i].decode())

	save_path += '.xlsx'
	wb.save(save_path)


if __name__ == '__main__':
	book_tag_lists = ['个人管理', '时间管理', '投资', '文化', '宗教']
	book_lists = do_spider(book_tag_lists)
	print_book_lists_excel(book_lists, book_tag_lists)